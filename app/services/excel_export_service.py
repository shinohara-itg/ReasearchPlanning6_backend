from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.excel_export import ExcelResearchItem


def _choices_to_text(choices: list[str]) -> str:
    return "\n".join([str(x).strip() for x in choices if str(x).strip()])


def _apply_common_style(ws, title: str, headers: list[str]) -> None:
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="E9F2F8")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[2].height = 28


def _style_body(ws, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)

    for row_idx in range(3, max_row + 1):
        ws.row_dimensions[row_idx].height = 44


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_screening_sheet(wb: Workbook, items: list[ExcelResearchItem]) -> None:
    ws = wb.active
    ws.title = "1_スクリーニング調査用"

    headers = [
        "No.",
        "設問項目",
        "設問タイプ",
        "選択肢例",
    ]
    _apply_common_style(ws, "1セット目：スクリーニング調査用", headers)

    for idx, item in enumerate(items, start=1):
        ws.append([
            item.number or idx,
            item.question,
            _question_type_display(item.question_type),
            _choices_to_text(item.choices_example),
        ])

    max_row = max(ws.max_row, 2)
    _style_body(ws, max_row, len(headers))
    _set_widths(ws, [8, 48, 16, 42, 20])
    ws.auto_filter.ref = f"A2:E{max_row}"


def _question_type_display(value: str) -> str:
    mapping = {
        "single": "SA",
        "multi": "MA",
        "numeric": "数値",
        "free_text": "FA",
        "single_grid": "SA表",
        "multi_grid": "MA表",
    }
    return mapping.get(str(value or "").strip(), value or "")


def _write_analysis_sheet(wb: Workbook, items: list[ExcelResearchItem]) -> None:
    ws = wb.create_sheet("2_本調査用")

    headers = [
        "No.",
        "SQ_ID",
        "設問項目",
        "設問タイプ",
        "選択肢例",
    ]
    _apply_common_style(ws, "2セット目：本調査用", headers)

    adopted_items = [
        item for item in items
        if not item.adoption_status or item.adoption_status == "adopted"
    ]

    for idx, item in enumerate(adopted_items, start=1):
        ws.append([
            item.number or idx,
            item.subq_id or "",
            item.question,
            _question_type_display(item.question_type),
            _choices_to_text(item.choices_example),
        ])

    max_row = max(ws.max_row, 2)
    _style_body(ws, max_row, len(headers))
    _set_widths(ws, [8, 14, 42, 16, 42])
    ws.auto_filter.ref = f"A2:E{max_row}"

def create_research_items_excel(
    screening_items: list[ExcelResearchItem],
    analysis_items: list[ExcelResearchItem],
) -> BytesIO:
    wb = Workbook()

    _write_screening_sheet(wb, screening_items)
    _write_analysis_sheet(wb, analysis_items)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output